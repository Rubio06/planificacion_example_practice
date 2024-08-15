import pandas as pd
import math
import json
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def calculate_ideal_size(days, last_sales_day, tvu, remove_product):
    if pd.isnull(last_sales_day) or isinstance(last_sales_day, pd._libs.tslibs.nattype.NaTType):
        return '0'
    
    return math.trunc((last_sales_day / days) * (tvu - remove_product))

def evaluate_masterpack(ddi_master_pack, tvu, stock, last_sales_day):
    if pd.isnull(ddi_master_pack) or ddi_master_pack == 'PRODUCTO SIN VENTA':
        return 'PRODUCTO SIN VENTA'
    
    try:
        ddi_master_pack = float(ddi_master_pack)
    except ValueError:
        return 'PRODUCTO SIN VENTA'
    
    if ddi_master_pack <= tvu:
        return 'OK'
    elif stock == 0 and last_sales_day == 0:
        return 'PRODUCTO SIN CARGA'
    elif stock > 0 and last_sales_day == 0:
        return 'PRODUCTO SIN VENTA'
    else:
        return 'BAJAR MASTERPACK'

def evaluate_packqty(ddi_packqty, tvu, stock, last_sales_day):
    if pd.isnull(ddi_packqty) or ddi_packqty == 'PRODUCTO SIN VENTA':
        return 'PRODUCTO SIN VENTA'
    
    try:
        ddi_packqty = float(ddi_packqty)
    except ValueError:
        return 'PRODUCTO SIN VENTA'
    
    if ddi_packqty <= tvu:
        return 'OK'
    elif stock == 0 and last_sales_day == 0:
        return 'PRODUCTO SIN CARGA'
    elif stock > 0 and last_sales_day == 0:
        return 'PRODUCTO SIN VENTA'
    else:
        return 'BAJAR PACKQTY'

def evaluate_innerpack(ddi_innerpack, tvu, stock, last_sales_day):
    if pd.isnull(ddi_innerpack) or ddi_innerpack == 'PRODUCTO SIN VENTA':
        return 'PRODUCTO SIN VENTA'
    
    try:
        ddi_innerpack = float(ddi_innerpack)
    except ValueError:
        return 'PRODUCTO SIN VENTA'
    
    if ddi_innerpack <= tvu:
        return 'OK'
    elif stock == 0 and last_sales_day == 0:
        return 'PRODUCTO SIN CARGA'
    elif stock > 0 and last_sales_day == 0:
        return 'PRODUCTO SIN VENTA'
    else:
        return 'BAJAR INNERPACK'

def calculate_and_validate(row):
    days = 28
    last_sales_day = row['VtaUltDiasCant']
    tvu = 60
    remove_product = 7
    ddi_master_pack = row['DDI_Masterpack']
    ddi_packqty = row['DDI_Packqty']
    ddi_innerpack = row['DDI_Innerpack']
    stock = row['Stock']
    
    ideal_size = calculate_ideal_size(days, last_sales_day, tvu, remove_product)
    validate_masterpack = evaluate_masterpack(ddi_master_pack, tvu, stock, last_sales_day)
    validate_packqty = evaluate_packqty(ddi_packqty, tvu, stock, last_sales_day)
    validate_innerpack = evaluate_innerpack(ddi_innerpack, tvu, stock, last_sales_day)
    
    results = {
        'validate_packqty': validate_packqty,
        'validate_innerpack': validate_innerpack,
        'validate_masterpack': validate_masterpack,
        'ideal_size': ideal_size
    }
    
    return results

while True:
    try: 
        results_list = []
        archive = input("Ingrese el nombre del archivo: ") + '.xlsx'
        sheet_name = input("Ingrese el nombre de la hoja: ")
        df = pd.read_excel('archivos/' + archive, sheet_name=sheet_name, header=1, index_col=None)
        break
    except Exception as e:
        print("----------------------")
        print("El archivo no existe o ha colocado texto vacío, intente de nuevo")
        print("----------------------")
        
for index, row in df.iterrows():
    try:
        results = calculate_and_validate(row)
        if not (results['validate_masterpack'] == 'PRODUCTO SIN VENTA' and 
                results['validate_packqty'] == 'PRODUCTO SIN VENTA' and 
                results['validate_innerpack'] == 'PRODUCTO SIN VENTA' and 
                results['ideal_size'] == '0'):
            results_list.append(results)
    except Exception as e:
        print(f"Error en la fila { index }: {e}")
# Convirtiendo a formato JSON, ident = 4 se utiliza para mejorar la legibilidad del JSON resultante
json_results = json.dumps(results_list, indent=4)
# Crear el directorio para guardar los resultados si no existe
output_dir = 'json_results'
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# Uniéndo las rutas para guardar el archivo result.json
output_file = os.path.join(output_dir, 'result.json')
with open(output_file, 'w') as f:
    f.write(json_results)
    
# Trabajando con el archivo Excel
df_excel = pd.DataFrame(results_list)
df_excel.columns = ['Valida_packqty', 'Valida_IP', 'Valida_MP', 'Tamaño ideal']
excel_file = os.path.join(output_dir, 'result.xlsx')

# Verificar si el archivo ya existe
if not os.path.exists(excel_file):
    df_excel.to_excel(excel_file, index=False)

    # Ajustar el ancho de las columnas en el archivo Excel
    wb = load_workbook(excel_file)
    ws = wb.active

    # Establecer anchos de columna específicos
    column_widths = {
        'A': 30,
        'B': 30,
        'C': 30,
        'D': 20
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    wb.save(excel_file)
    wb.close()

    print(f"Resultados guardados en {output_file}")
    print(f"Resultados sobre archivo Excel se han guardado correctamente en {excel_file}")
else:
    print("El archivo ya existe, intente nuevamente")