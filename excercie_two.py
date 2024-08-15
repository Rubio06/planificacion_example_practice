import pandas as pd
import os
from openpyxl import load_workbook

# Ruta del archivo
path = 'archivos'
fillname = input('Ingrese el nombre del archivo: ') + '.xlsx'
fullpath = os.path.join(path, fillname)

# Leer el archivo Excel con pandas para la tabla
df = pd.read_excel(fullpath, sheet_name='Validación', header=1)

# Imprimir los nombres de las columnas para verificar
print("Columnas del DataFrame:", df.columns)

# Leer el archivo Excel con openpyxl para acceder a una celda específica
wb = load_workbook(fullpath)
sheet = wb['Validación']

print(sheet)

# Especificar la celda que quieres leer (fila 1 y columna AB)
fila = 1
columna = 'AB'
celda = f'{columna}{fila}'
dato_celda = sheet[celda].value
# print(f"Dato en la celda {celda}: {dato_celda}")

# Iterar sobre las filas del DataFrame
if 'TVU' in df.columns:
    for index, row in df.iterrows():
        print(type(row["TVU"]))
else:
    print("La columna 'tvu' no existe en el DataFrame.")
    

# Imprimir el DataFrame
print(df)

# wb = load_workbook(fullpath)
# sheet = wb[sheet_name]
# row = 1
# column = 'T'
# cell = f"{column}{row}"
# data_cell = sheet[cell].value
# print(f"Dato en la celda {cell}: {data_cell}")

# if data_cell in days:
#     found = True
